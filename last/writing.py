import openpyxl

file = "D:\empty.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook['Sheet2']

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value = "Welcome"

sheet.cell(1,1).value = "BookName"
sheet.cell(1,2).value = "Product Name"
sheet.cell(1,1).value = "BookName"

workbook.save(file)

