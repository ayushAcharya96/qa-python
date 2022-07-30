import os
import openpyxl

import xlUtils

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
location = os.getcwd()

# def chrome_setup():
#     service = Service()
#     driver = webdriver.Chrome(service= service)
#     return driver

# driver = chrome_setup()
# driver.implicitly_wait(10)

# file -> workbook -> sheets -> cells
file = "C:\BookName.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active #(get active sheet if only one sheet)
# sheet = workbook["Sheet1"]
rows = sheet.max_row
cols = sheet.max_column
# xlUtils.readData(file, sheet,)


print(sheet.cell(2,1).value)
print(sheet.cell(5,3).value)
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r,c).value, end='      ')
    print()
